import glob
import os
import tabula
import datetime
from openpyxl import load_workbook


class Check_download_file:
    def __init__(self):
        self.file_path = "C:/Users/morris.lin/Downloads"
        self.file_time = datetime.datetime.now().strftime('%d_%m_%Y')
        self.pdf_path = rf'C:/Users/morris.lin/Downloads/Cname_Test_{self.file_time}.pdf'
        self.excel_path = rf'C:/Users/morris.lin/Downloads/Cname_Test_{self.file_time}.xlsx'

    def check_file_exist(self):
        try:
            pdfs = str(glob.glob("{}/*.pdf".format(self.file_path)))
            excels = str(glob.glob("{}/*.xlsx".format(self.file_path)))
            # print(pdfs)
            # 確認 pdf存在
            assert self.file_time in pdfs
            # 確認excel 存在
            assert self.file_time in excels
        except Exception as msg:
            print('error:', msg)

    def open_pdf(self):
        # pip3 install tabula-py
        # pip3 install pandas
        try:
            df = tabula.read_pdf(self.pdf_path, pages='all')

            pdf_content = []

            # 整理顯示資料  dataframe index
            for index, row in df[1].iterrows():
                for value in row:
                    # print(value)
                    # print("value: {}".format(value))
                    pdf_content.append(value)

            # check default data
            assert 'https://www.example.com' in str(df[0])
            assert '風險' in str(df[1])
            assert 'A1:Injection' and 'A2:Broken Authentication' and 'CWE 91:' \
                   and 'XML Injection (aka Blind XPath Injection)' in str(df)
            assert len(df) > 2 and len(df[1]) > 1 and len(df[2]) > 1
            # check data isn't 'null'
            assert len(pdf_content[1]) > 5 and len(pdf_content[1]) > 1
        except Exception as msg:
            print('error:', msg)

    def open_excel(self):
        # pip install openpyxl
        try:
            excel = load_workbook(self.excel_path)
            #check data isn't 'null'
            assert 'Issues' and 'URLs' in excel.sheetnames
            assert len(excel['Issues']['B2'].value) and len(excel['Issues']['C2'].value) and \
                   len(excel['Issues']['D2'].value) and len(excel['Issues']['E2'].value) and \
                   len(excel['Issues']['F2'].value) and len(excel['Issues']['G2'].value) > 1
            assert len(excel['URLs']['B2'].value) and len(excel['URLs']['C2'].value) and len(excel['URLs']['D2'].value) > 1
        except Exception as msg:
            print('error:', msg)




